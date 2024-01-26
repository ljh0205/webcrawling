
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from subprocess import CREATE_NO_WINDOW
serve=Service()
serve.creation_flags=CREATE_NO_WINDOW
driver=webdriver.Chrome(service=serve)


import time, datetime, os #datetime==시간
import openpyxl #openpyxl==py<=>xl

now=str(datetime.datetime.now())[ : 16] #datetime 1번째 글자에서 15번째 글자까지 자르기
folderName=now.replace(":","_") #   :  ==> _
os.mkdir(folderName)  # 폴더를 만들어서 foldername 괄호로 설정

kw=['2024년', '정치']

wb=openpyxl.Workbook() # xl 워크북 만들기

for i in range(len(kw)):  #키워드의 길이 만큼 밑 코드 반복
    ws=wb.create_sheet() #xl worksheet 생성
    ws.title=kw[i] #워크쉬트의 제목을 kw리스트의 i번째 인덱스로 설정 
    ws.column_dimensions['A'].width=90 #A열 넓이 설정
    
    url = "https://search.naver.com/search.naver?where=news&ie=utf8&sm=nws_hty&query=" + kw[i] #url + 키워드로 링크 설정 {url + 키워드 == 키워드를 검색한 링크}
    driver.get(url)
    time.sleep(2)

    list_news = driver.find_element("class name", "list_news")
    news_boxes = list_news.find_elements("class name", "bx")

    for j in range(len(news_boxes)):  #news_boxes의 갯수 만큼 밑 내용 반복
        driver.execute_script("arguments[0].scrollIntoView(true);", news_boxes[j])  #뉴스 카드 개체 하나의 전체 부분이 나올때까지 스크롤 하기 
        file = f"{folderName}/{i+1}_{kw[i]}-{j+1}.png"  #파일 이름 설정
        news_boxes[j].screenshot(file)  #news_boxes를 스크린 샷
        
        ws.row_dimensions[j+1].height = 100 #행의 높이 설정
        img = openpyxl.drawing.image.Image(file)   #xl에 넣을 수 있는 사진 파일로 변경
        ws.add_image(img, f'A{j+1}')  #이미지를 a열 j + 1행에 저장

        title = news_boxes[j].find_element("class name", "news_tit")  
        print(j+1, title.text) #뉴스 타이틀 출력

        link = title.get_attribute("href") #링크 저장
        ws[f'B{j+1}'].value = "기사링크" #기사링크 b열의 j + 1행에 *'기사링크' 넣기 
        ws[f'B{j+1}'].hyperlink = link#하이퍼링크를 *기사링크에 걸기

    print()
    
wb.remove(wb["Sheet"])  #필요 없는 워크 쉬트 삭제
wb.save(f"{folderName}/{folderName}_결과.xlsx")  #워크북 저장

